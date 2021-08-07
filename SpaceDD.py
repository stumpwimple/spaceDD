#########################################################################################
#########################################################################################
#########################################################################################
#########################################################################################
###########################                                           ###################
###########################    SPACE D&D RANDOM UNIVERSE GENERATOR    ###################
###########################    WORKING TITLE: TAXONOMY                ###################
#########################################################################################
#########################################################################################
#########################################################################################
#########################################################################################
#########################################################################################





import random
import math
import copy
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import OrderedDict
from tkinter import *
#from tkinter.ttk import *
import tkinter.font as font
from tkinter import messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
import webbrowser

alphanum = 'abcdefghijklmnopqrstuvwxyz0123456789'
basic_region_type = ('Mountain','Island','Forest','Plains','Swamp','Ocean','Desert')
basic_planet_type = ('Hot Planet','Habitable Planet','Gas Planet','Icy Gas Planet')
basic_weather = ('Sunny','Clear sky','Overcast','Wind','Precipitation')
lifeform_type_list = ('Plant-like Lifeform','Advanced Lifeform','Hybrid Lifeform', 'Chimeric Lifeform', 'Monster Manual Lifeform','Mythological Lifeform')
planet_water_type = ('frozen caps', 'One Sea','Two Seas','ocean 25%', 'ocean 50%', 'ocean 75%', 'ocean 100%')
dice_type = (4,6,8,10,12,20)

screenX = 1600
screenY = 900

#Initializing Dictionary of lists and list categories
list_of = {}
categories_of_list = {}

min_planets = 3
max_planets = 7 #9

max_planet_size = 20

difficulty = 12   # Max DC anything will roll



#########################################################################################
######  FUNCTIONS  ######################################################################
#########################################################################################

#creates random alphanumeric string of length num
def rand_string(num):
    return_string = ''
    for i in range(num):
        return_string += random.choice(alphanum)
    return return_string

def rando(itemList):
    print(sr(random.choice(list_of[itemList])))


#returns num "tabs" each tab equals 3 spaces
def tab(num):
    return "   "*num

#imports excel workbook, creates random generation lists, and list of random list categories. Also collects a list of categories this list could belong to and stores it in a list. 
def import_random_lists_from_file():
    wb = load_workbook(filename='Random Lists.xlsx')
    ws = wb["LISTS"]

    for col in ws.iter_cols():
        this_list = []
        list_categories = []
        for cell in col:    
            #stores list_name for this list of random items       
            if cell.row == 1 and cell.value:
                list_name = cell.value
            #stores categories of list and stores them. Excel file uses rows 2-10 as categories
            elif (1 < cell.row <= 10) and cell.value:
                list_categories.append(cell.value)
            #stores items in list of random items
            elif cell.value:
                this_list.append(cell.value)
            else:
                break
        list_of[list_name]=this_list
        categories_of_list[list_name]=list_categories

#Sub-Resolves any nested Random Lists
def sr(this_str): 
    while ("[" in this_str):
        nested_list = this_str[this_str.find("[")+1:this_str.find("]")]        
        this_result = random.choice(list_of[nested_list])
        try:
            this_result = str(this_result)
        except ValueError:
            pass
        this_str = this_str.replace ("["+nested_list+"]","|"+this_result+"|")
    return this_str

#Returns True/False based on num % chance
def chance(num):
    return (True if (random.randint(0,99) < num) else False)

#Returns Lengthened string by appending spaces at the end of it. Used to fix formatting issues in tKinter
def str_len(new_string_length,the_string):
    length = len(the_string)
    while length < new_string_length:
        the_string += " "            
        length = len(the_string)
    return the_string

#function to check isThisInt?
def isThisInt(s):
    try: 
        int(s)
        return True
    except ValueError:
        return False


#########################################################################################
######  CLASSES  ########################################################################
#########################################################################################

class DefaultClassTemplate():
    
    def __init__(self):
        pass
    
    def __str__(self):
        pass
    
    def generate(self,opt1=0,opt2=0):
        pass
#########################################################################################
##  class Trait  ########################################################################
######################################################################################### 

#Trait Class - A class that displays a trait ui, with a number of other ui elements meant for running a pprpg 
class Trait():
    
    def __init__(self,trait_name,trait_formula,trait_proficiency_formula="random.randint(1,difficulty)",summary_trait=False,trait_unit=""):
        self.identity=trait_name
        self.formula=trait_formula
        self.unit=trait_unit
        self.value = eval(self.formula)
        self.proficiency_formula = str(trait_proficiency_formula)
        self.proficiency = eval(self.proficiency_formula)
        self.summary_trait=summary_trait
        self.is_discovered = False
    
    def __str__(self):
        trait_str = str_len(25,str(self.identity) + ":")
        trait_str += str_len(60,str(self.value) + " " +str(self.unit))[:60] # print only the left 60 charactes of value
        trait_str += str(self.proficiency)
        return trait_str
    
    def generate(self,trait_name=0):
        pass

    def reroll(self):
        self.value=eval(self.formula)
        self.proficiency = eval(self.proficiency_formula)

    # def roll_proficiency(self):
    #     print("You rolled a " + str(random.randint(1,self.proficiency)))

    


    def generate_trait_ui(self,trait_frame,ui_row):
        def update_discovered():
            if discovered.get() == True:
                self.is_discovered=True
            else:
                self.is_discovered=False 

        def update_proficiency(prof_value):
            if isThisInt(prof_value.get()):
                self.proficiency=int(prof_value.get())
            elif prof_value.get() != "":
                self.ui_proficiency_value.set(self.proficiency)

        def update_value(trait_value):
            self.value=trait_value.get()



        discovered = BooleanVar()
        if self.is_discovered == True: discovered.set(True)
        ui_discovered = Checkbutton(trait_frame, variable = discovered, command = update_discovered )
        ui_discovered.grid(row=ui_row,column=0,padx=1,sticky=W)

        self.ui_proficiency_value = StringVar()
        self.ui_proficiency_value.set(self.proficiency)
        self.ui_proficiency_value.trace("w",lambda name,index,mode, prof_value=self.ui_proficiency_value: update_proficiency(prof_value))
        ui_proficiency = Entry(trait_frame,width = 3,textvariable = self.ui_proficiency_value)
        ui_proficiency.grid(row=ui_row,column=1,padx=1,sticky=W)

        ui_roll_button = Button(trait_frame,image=image_d6,height=13,width=13, command = lambda self = self: roll_proficiency(self))
        ui_roll_button.grid(row=ui_row,column=2,padx=1,sticky=W)

        ui_trait_identity = Label(trait_frame, text = str(self.identity))
        ui_trait_identity.grid(row=ui_row,column=3,padx=1,sticky=W)

        self.ui_value = StringVar()
        self.ui_value.set(str(self.value)+" "+self.unit) 
        self.ui_value.trace("w",lambda name,index,mode, trait_value=self.ui_value: update_value(trait_value))
        ui_trait_value = Entry(trait_frame,width = 40,textvariable=self.ui_value)
        ui_trait_value.grid(row=ui_row,column=4,padx=1,sticky=W)

        ui_lookup_trait_button = Button(trait_frame,image=image_lookup,height=13,width=13,command = lambda self = self: lookup_link(self))
        ui_lookup_trait_button.grid(row=ui_row,column=5,padx=1,sticky=W)

        ui_reroll_trait_button = Button(trait_frame,image=image_reroll,height=13,width=13,command = lambda self = self: reroll_trait(self)) 
        ui_reroll_trait_button.grid(row=ui_row,column=6,padx=1,sticky=W)

        ui_journal_button = Button(trait_frame,image=image_journal,height=13,width=13,command=lambda: test_button("You pressed Journal"))
        ui_journal_button.grid(row=ui_row,column=7,padx=1,sticky=W)
       
        #ui_notes[0]= Entry(traitFrame,width = 20)
        #ui_notes[0].insert(0,"Adventure Notes - replace with dialog pop-up and logging to journal")
        #ui_notes[0].grid(row=0,column=7,padx=1,sticky=W)

##########################################
##  class Sol  ###########################
##########################################
    
#Solar System Class
class Sol():
    
    def __init__(self,sol_type=0,parent="self"):
        self.parent = parent
        self.object_type = "Sol" #(Sol)ar System
        self.trait = OrderedDict()
        self.subobject = []

        if(sol_type==0):
            self.trait['identity']= Trait('Identity',"'SOL-' + rand_string(3)")
            self.trait['sol_type']= Trait('Solar System type',"'Solar - One Star'",summary_trait=True)
            self.trait['star_temp']= Trait('Star Temperature',"random.randint(20,100)*100",trait_unit="K")
            self.trait['star_diameter']= Trait('Star Diameter',"round("+str(self.trait['star_temp'].value)+"*0.000245,3)",trait_unit="Mkm")
            self.trait['edge_helio']= Trait('Edge of Heliosphere',"("+str(self.trait['star_temp'].value)+" // 115)**2",trait_unit="Mkm",summary_trait=True)
            self.trait['edge_gas_zone']= Trait('Edge of Gas Zone',str(self.trait['edge_helio'].value)+" // 2",trait_unit="Mkm")
            self.trait['edge_hab_zone']= Trait('Edge of Hab Zone',str(self.trait['edge_helio'].value)+"// random.randint(3,6)",trait_unit="Mkm")
            self.trait['edge_hot_zone']= Trait('Edge of Hot Zone',str(self.trait['edge_helio'].value)+" // random.randint(7,9)",trait_unit="Mkm")
            self.trait['effect_chance'] = Trait("Effect Chance","random.randint(1,50)")
            if chance(self.trait['effect_chance'].value): self.trait['system_effect'] = Trait('System Effect',"sr(random.choice(list_of['SpaceStoryHook']))",summary_trait=False)
         
            self.trait['num_planets'] = Trait("Number of planets","random.randint(min_planets,max_planets)",summary_trait=True,trait_unit="Planets")
            for i in range(1,self.trait['num_planets'].value+1):
                orbit = int(random.random()*self.trait['edge_gas_zone'].value + 15*self.trait['star_diameter'].value)
                planet_type = self.get_planet_type_by_orbit(orbit)                            
                this_planet = Planet(planet_type,orbit,i,parent=self)
                self.subobject.append(this_planet)
            
            #print(type(self.subobject[0]))

    def __str__(self):
        sol_str=""
        for tr in self.trait:
            sol_str += tab(1) + str(self.trait[tr])
            sol_str += "\n" 
        sol_str += "\n"            

        sol_str += tab(1) + f"---PLANETS ({self.trait['num_planets'].value})---\n"
        for so in self.subobject:
            sol_str += str(so)      
            sol_str += "\n" 
        sol_str += "\n" 
            
        return sol_str
            
    def get_planet_type_by_orbit(self, orbit):
        if orbit < self.trait['edge_hot_zone'].value:
            return 'Hot Planet'
        elif orbit < self.trait['edge_hab_zone'].value:
            return 'Habitable Planet'
        elif orbit < self.trait['edge_gas_zone'].value:
            return 'Gas Planet'
        else:
            return 'Mysterious Planet'



##########################################
##  class Planet  ########################
##########################################
   
#Planet Class    
class Planet():
    
    def __init__(self,planet_type=0,orbit=0,planet_id=0,parent="self"):        
        self.parent = parent
        self.object_type = "Planet"
        self.trait = OrderedDict()
        self.trait['identity'] = Trait("Identity","rand_string(4)")
        self.trait['size']=Trait("Planet Size","roll_size()",trait_unit="Regions",summary_trait=True)  
        self.trait['orbit_radius'] = Trait('Orbit',"'"+str(orbit)+"'",trait_unit="Mkm")
        self.trait['orbit_radians'] = Trait('Orbit',"2*math.pi*random.random()",trait_unit="Radians")
        self.trait['color'] = Trait("Color from Space","sr(random.choice(list_of['PythonColors']))")
        self.trait['planet_type'] = Trait('Planet Type',"'"+str(planet_type)+"'",summary_trait=True)
        self.trait['chance_of_water'] = Trait("Chance of Water","'0'")
        self.trait['planet_water']=Trait("Planet water cover","'random.choice(planet_water_type)'",summary_trait=True)
        self.trait['planet_water'].value = "None"
        self.trait['gravity'] = Trait("Gravity","sr(random.choice(list_of['Gravity']))",summary_trait=True)
        self.trait['breathability'] = Trait("Breathability","sr(random.choice(list_of['Breathability']))") 
        self.trait['topography'] = Trait('Topography',"'Rocky Topography'",summary_trait=True)
        self.trait['chance_of_life'] = Trait("Chance of Life","'0'")        
        self.trait['chance_of_sentient_life'] = Trait("Chance of Sentient Life","'0'")
        self.trait['effect_chance'] = Trait("Effect Chance","random.randint(1,50)",summary_trait=True) 
        if chance(self.trait['effect_chance'].value): self.trait['effect'] = Trait("Planet Effect","sr(random.choice(list_of['PlanetEffect']))") 

        self.trait['planet_id'] = Trait('Planet Id',"'"+str(planet_id)+"'")

        #Declare empty subobject and lifeform lists
        self.region = []
        self.lifeform = []
        self.sentient = []

        self.subobject = []
        
        #If planet is unassigned on creation, randomly generates from basic_planet_type       
        if(self.trait['planet_type'].value==0):            
            self.trait['planet_type'] = Trait("Planet Type","random.choice(basic_planet_type)")
        
        #Give Planet type based traits
        if self.trait['planet_type'].value == 'Hot Planet':
            self.trait['identity'].value = "'HOT-"+rand_string(4)+"'"
            self.trait['identity'].formula = "'HOT-"+rand_string(4)+"'"
            self.trait['avg_temp'] = Trait("Global Climate","sr(random.choice(list_of['HotAvgTemp']))")
            self.trait['chance_of_water'].value = 20
            self.trait['chance_of_life'].value = 20
            self.trait['chance_of_sentient_life'].value = 2
        elif self.trait['planet_type'].value == 'Habitable Planet':
            self.trait['identity'].value = "'HAB-"+rand_string(4)+"'"
            self.trait['identity'].formula = "'HAB-"+rand_string(4)+"'"
            self.trait['avg_temp'] = Trait("Global Climate","sr(random.choice(list_of['HabAvgTemp']))")
            self.trait['chance_of_water'].value = 40
            self.trait['chance_of_life'].value = 40
            self.trait['chance_of_sentient_life'].value = 4   
        elif self.trait['planet_type'].value == 'Gas Planet':
            self.trait['identity'].value = "'GAS-"+rand_string(4)+"'"
            self.trait['identity'].formula = "'GAS-"+rand_string(4)+"'"
            self.trait['avg_temp'] = Trait("Global Climate","sr(random.choice(list_of['GasAvgTemp']))") 
            self.trait['breathability'].value = 'Not breathable'
            self.trait['chance_of_water'].value = 0
            self.trait['chance_of_life'].value = 5
            self.trait['chance_of_sentient_life'].value = 1
        elif self.trait['planet_type'].value == 'Icy Gas Planet':
            self.trait['identity'].value = "'ICE-"+rand_string(4)+"'"
            self.trait['identity'].formula = "'ICE-"+rand_string(4)+"'"
            self.trait['avg_temp'] = Trait("Global Climate","sr(random.choice(list_of['IceGasAvgTemp']))")
            self.trait['breathability'].value = 'Not breathable'
            self.trait['chance_of_water'].value = 0
            self.trait['chance_of_life'].value = 5
            self.trait['chance_of_sentient_life'].value = 1
        elif self.trait['planet_type'].value == 'Mysterious Planet':
            self.trait['identity'].value = "'UPO-"+rand_string(4)+"'"
            self.trait['identity'].formula = "'UPO-"+rand_string(4)+"'"
            self.trait['avg_temp'] = Trait("Global Climate","sr(random.choice(list_of['GasAvgTemp']))")
            self.trait['chance_of_water'].value = 50
            self.trait['chance_of_life'].value = 50
            self.trait['chance_of_sentient_life'].value = 5
        else:
            self.trait['chance_of_water'].value = 75
            self.trait['chance_of_life'].value = 75
            self.trait['chance_of_sentient_life'].value = 75
            self.trait['planet_type'] = Trait('Planet Type',"random.choice(basic_planet_type)")
            print(f"Planet {self.trait['planet_id'].value} did not have a standard planet_type")
            self.trait['identity'].value = "'UNKO-"+rand_string(4)+"'"
            self.trait['identity'].formula = "'UNKO-"+rand_string(4)+"'"
            self.trait['avg_temp'] = Trait("Global Climate","sr(random.choice(list_of['GasAvgTemp']))")       
        
        #print(f"planet {self.planet_id} is {self.planet_type} and has {self.chance_of_water} chance of water")
         
        #Check if Planet has water
        if random.randint(0,100) < self.trait['chance_of_water'].value:
            self.trait['planet_water'].value = random.choice(planet_water_type)
            self.trait['chance_of_life'].value += self.trait['chance_of_water'].value/2
            self.trait['topography'].value = 'Rocky-Water Topography'
    
        #Generate Regions of Planet based self.size
        for i in range(1,self.trait['size'].value+1):
            this_region = Region(self,0,i,parent=self)
            self.region.append(this_region)



        #Lifeforms are generated by planet, but later populated into regions with regional traits.
        #Generate Lifeforms (if any)
        if True: #random.randint(0,100) < self.trait['chance_of_life'].value:
            for i in range(1,random.randint(4,self.trait['size'].value)):
                this_lifeform = Lifeform(parent=self)
                self.lifeform.append(this_lifeform)
            self.region = self.populate_lifeforms(self.region,self.lifeform)            
        
        #Sentient Lifeforms are generated by planet, but later populated into regions with regional traits.    
        #Generate Sentient Life (if any)
        if True: #random.randint(0,100) < self.trait['chance_of_sentient_life'].value:
            for i in range(1,max(2,random.randint(-3,4))):
                print("Generating a Sentient")
                this_sentient = SentientLife(parent=self)
                self.sentient.append(this_sentient)
                
            self.region = self.populate_sentients(self.region,self.sentient)

        self.subobject = self.region  
    
    def __str__(self):        

        planet_str=""
        for tr in self.trait:
            planet_str += tab(2) + str(self.trait[tr]) + "\n" 

        planet_str += "\n"       
        
        if self.sentient:
            planet_str += tab(2) +  f"---SENTIENT SPECIES ({len(self.sentient)})---\n"
            for s in self.sentient:
                planet_str += s.sentient_summary()
            
        planet_str += tab(2) +  f"---REGIONS ({self.trait['size'].value})---\n"
        for so in self.subobject:
            planet_str += str(so)
            planet_str += "\n" 
        return planet_str

    #Populate lifeforms into regions
    def populate_lifeforms(self,region,lifeform):
        for r in region:        
            r.subobject.extend(copy.deepcopy(random.sample(lifeform,random.randint(0,len(lifeform)))))
            for so in r.subobject:
                if so.object_type == "Lifeform":              
                    so.parent = r
                    if so.trait['lifeform_type'] == "Plant-like Lifeform":                        
                        so.trait['regional_trait'] = Trait('Region  Trait', "sr(random.choice(list_of['PlantFeatures']))")
                    else:
                        so.trait['regional_trait'] = Trait('Regional Trait', "sr(random.choice(list_of['AdvLifeformUniques']))")
        return region
   
    #Populate Sentient lifeforms into regions 
    def populate_sentients(self,region,sentient):
        for r in region:        
            r.subobject.extend(copy.deepcopy(random.sample(sentient,random.randint(0,len(sentient)))))
            for so in r.subobject:
                if so.object_type == "Sentient":              
                    so.parent = r

        return region

#returns a random planet size, rounds to "nearest" dice size (d4,d6,d8,d10,d12)
def roll_size():
    size_roll = random.randint(1,max_planet_size+1)
    if size_roll > 18:
        return 20
    elif size_roll > 12:
        return 12
    elif size_roll <= 4:
        return 4
    elif size_roll % 2 == 1:
        return size_roll-1
    else:
        return size_roll

##########################################
##  class Region  ########################
##########################################


#Region Class - A planet has 4-20 regions, each with unique topography, weather, and wildlife. 
class Region():
    
    def __init__(self,parent_planet=0,region_type=0,region_id=0, parent = "self"):
        self.parent = parent
        self.object_type = "Region"
        self.trait = OrderedDict()
        self.parent_planet = parent_planet
       # print(planet)
        self.region_type = region_type
        self.region_id=region_id
        self.trait['identity'] = Trait("Identity","'REGION " + str(self.region_id) + "'")
        
        if self.parent_planet.trait['topography'].value =='Rocky Topography':
            self.trait['terrain'] = Trait('Terrain',"sr(random.choice(list_of['RockyTopography']))",summary_trait=True)
        elif self.parent_planet.trait['topography'].value =='Rocky-Water Topography':
            self.trait['terrain'] = Trait('Terrain',"sr(random.choice(list_of['RockyWaterTopography']))",summary_trait=True)
        elif self.parent_planet.trait['topography'].value =='Water Topography':
            self.trait['terrain'] = Trait('Terrain',"sr(random.choice(list_of['WaterTopography']))",summary_trait=True)   
        
        
        if self.parent_planet.trait['avg_temp'].value.split()[0] == 'Hot':
            self.trait['region_weather'] = Trait('Regional weather',"sr(random.choice(list_of['HotRegionTemps'])) + ', ' + sr(random.choice(list_of['Wind']))",summary_trait=True)
        elif self.parent_planet.trait['avg_temp'].value.split()[0] == 'Moderate':
            self.trait['region_weather'] = Trait('Regional weather',"sr(random.choice(list_of['ModerateRegionTemps'])) + ', ' + sr(random.choice(list_of['Wind']))",summary_trait=True)
        elif self.parent_planet.trait['avg_temp'].value.split()[0] == 'Cold':
            self.trait['region_weather'] = Trait('Regional weather', "sr(random.choice(list_of['ColdRegionTemps'])) + ', ' + sr(random.choice(list_of['Wind']))",summary_trait=True)
    

        
        self.subobject = []

        for r in range(1,max(2,random.randint(-2,3))):
            this_resource = Resource(parent=self)
            self.subobject.append(this_resource)

        self.lifeform = []
        self.sentient = []

    def __str__(self):    

        region_str=""
        for tr in self.trait:
            region_str += tab(3) + str(self.trait[tr]) + "\n" 

        region_str += "\n" 

        region_str +=  tab(3) + "---RESOURCES---\n"    
        for so in self.subobject:
            if so.object_type=="Resource":
                region_str += str(so) + "\n" 

        region_str += "\n"      
        
        if self.lifeform:
            region_str +=  tab(3) + "---LIFEFORMS---\n"
            for so in self.subobject:
                if so.object_type == "Lifeform":
                    region_str += str(so) + "\n" 
                    region_str += "\n"

        if self.sentient:
            region_str +=  tab(3) + "---SENTIENT LIFE---\n"  
            for so in self.subobject:
                if so.object_type=="Sentient":
                    region_str += str(so) + "\n" 
        
        region_str += "\n"  

        return region_str

##########################################
##  class Resource  ######################
##########################################
    
#Resource Class - Generates resource, mineral or gas, possibly future implementation with liquids
class Resource():
    
    def __init__(self,resource_type='',parent="self"):
        self.parent = parent
        self.object_type = "Resource"
        self.resource_type=resource_type

        self.trait = OrderedDict() 
        self.trait['identity'] = Trait("Catalog #","' '")
        if self.resource_type == '': 
            self.trait['resource_type'] = Trait("Resource Type","random.choice(['Mineral','Mineral','Gas'])",summary_trait=True)
        elif self.resource_type != '': 
            self.trait['resource_type'] = Trait(str(self.resource_type),"random.choice(['Mineral','Mineral','Gas'])",summary_trait=True)

        if self.trait['resource_type'].value == 'Mineral':
            self.trait['identity'] = Trait("Catalog #","'min-' + rand_string(3)")
            self.trait['description'] = Trait("Resource Description","sr(random.choice(list_of['ColorCommon'])) + ' ' + sr(random.choice(list_of['MineralDescription']))",summary_trait=True)
            if chance(50): self.trait['uses'] = Trait("Potential uses","sr(random.choice(list_of['MineralUses']))",summary_trait=True)
                                             
        elif self.trait['resource_type'].value == 'Gas':
            self.trait['identity'] = Trait("Catalog #","'gas-' + rand_string(3)")
            self.trait['description'] = Trait("Resource Description","get_gas_color() + ' ' + sr(random.choice(list_of['GasDescription']))",summary_trait=True)
            if chance(50): self.trait['uses'] = Trait("Potential uses","sr(random.choice(list_of['GasUses']))",summary_trait=True)            
    
    def __str__(self): 
        resource_str=""
        for tr in self.trait:
            resource_str += tab(4) + str(self.trait[tr]) + "\n" 
        resource_str += "\n"  
        return resource_str    

#Generate random color of gas resource, 90% chance invisible
def get_gas_color():
    if chance(10):
        color = sr(random.choice(list_of['ColorCommon']))
    else:
        color = 'Invisible'
    return color
    
##########################################
##  class Item           #################
##########################################
            
    
#Future Item Class - Items, Loot, Armor, Weapons!    
class Item():

    def __init__(self):
        pass
    
    def __str__(self):
        pass
    
    def generate(self, opt1=0,opt2=0):
        pass


##########################################
##  class Lifeform  ######################
##########################################
    
#Function to Generate Lifeform Type
def get_lifeform_type():
    roll = random.randint(0,100)
    
    if roll < 50:
        lifeform_type = 'Plant-like Lifeform'
    elif roll <80:
        lifeform_type = 'Advanced Lifeform'
    elif roll <85:
        lifeform_type = 'Advanced Aquatic Lifeform'
    elif roll <90:
        lifeform_type = 'Hybrid Lifeform'
    elif roll <94:
        lifeform_type = 'Chimeric Lifeform'
    elif roll <98:
        lifeform_type = 'Mythological Lifeform'
    elif roll <99:
        lifeform_type = 'Monster Manual Lifeform'                
    else:
        lifeform_type = random.choice(lifeform_type_list)
    return lifeform_type

    
#Lifeform Class - Generates lifeform using various types as templates 
class Lifeform():
    
    def __init__(self,lifeform_type=0, parent="self"):

        self.parent = parent
        self.object_type = "Lifeform"
        self.trait = OrderedDict()
        self.trait['identity'] = Trait("Identity","'LIF-'+rand_string(3)")
        
        if lifeform_type == 0:
            self.trait['lifeform_type']=Trait("Lifeform Type","get_lifeform_type()")
        else:
            self.trait['lifeform_type']=Trait("Lifeform Type","'"+lifeform_type+"'")

        self.trait['species']=Trait("Species","sr(random.choice(list_of['AdvancedLifeforms']))",summary_trait=True)
        self.trait['colors'] = Trait("Lifeform Colors","sr(random.choice(list_of['ColorCommon'])) + '/' + sr(random.choice(list_of['ColorCommon']))",summary_trait=True)
        self.trait['size'] = Trait("Lifeform Size","sr(random.choice(list_of['AdvLifeformSize']))",summary_trait=True)
        self.trait['density'] = Trait("Lifeform Density","sr(random.choice(list_of['AdvLifeformDensity']))",summary_trait=True)
        self.trait['diet'] = Trait("Lifeform Diet","sr(random.choice(list_of['AdvLifeformDiet']))")
        self.trait['storyhook'] = Trait("Story Hook","sr(random.choice(list_of['LifeformStoryHook']))")
        self.trait['reproduction'] = Trait("Reproduction Method","sr(random.choice(list_of['AdvLifeformRepro']))")
        self.trait['uses'] = Trait("Lifeform Uses", "sr(random.choice(list_of['AdvLifeformUses']))")
        self.trait['demeanor'] = Trait("Lifeform Demeanor","sr(random.choice(list_of['AdvLifeformDemeanor']))")
        self.trait['combat_profile'] = Trait("Combat Profile","sr(random.choice(list_of['BaseLifeformCombatProfile']))")
        self.trait['reaction'] = Trait("Reaction to...","sr(random.choice(list_of['AdvLifeformReaction']))") # Build system where this is randomly generated upon player action

        num_uniques = 0

        #Determines Base species based on lifeform_type
        if self.trait['lifeform_type'].value == 'Advanced Lifeform':
            self.trait['species'] = Trait("Adv Lifeform Species","sr(random.choice(list_of['AdvancedLifeforms']))",summary_trait=True)
            num_uniques = max(2,random.randint(-1,3))

        elif self.trait['lifeform_type'].value == 'Hybrid Lifeform':
            self.trait['species'] = Trait("Hybrid Species","'Half ' + sr(random.choice(list_of['AdvancedLifeforms'])) + \
                                                    ', half ' + sr(random.choice(list_of['AdvancedLifeforms']))",summary_trait=True)
            num_uniques = max(2,random.randint(-1,3))
            
        elif self.trait['lifeform_type'].value == 'Chimeric Lifeform':
            self.trait['species'] = Trait("Chimeric Species","'Part ' + sr(random.choice(list_of['AdvancedLifeforms'])) + \
                                                    ', part ' + sr(random.choice(list_of['AdvancedLifeforms'])) + \
                                                    ', part ' + sr(random.choice(list_of['AdvancedLifeforms']))",summary_trait=True)
            num_uniques = max(2,random.randint(-1,3))              
                
        elif self.trait['lifeform_type'].value == 'Mythological Lifeform':
            self.trait['species'] = Trait("Mythological Species","sr(random.choice(list_of['MythologicalLifeforms']))",summary_trait=True)
            num_uniques = max(2,random.randint(-1,3))
            
        elif self.trait['lifeform_type'].value == 'Advanced Aquatic Lifeform':
            self.trait['species'] = Trait("Aquatic Species","sr(random.choice(list_of['AdvancedLifeformsAquatic']))",summary_trait=True)
            num_uniques = max(2,random.randint(-1,3))

        elif self.trait['lifeform_type'].value == 'Monster Manual Lifeform':
            self.trait['species'] = Trait("MM Species","sr(random.choice(list_of['MMLifeforms']))",summary_trait=True)
            num_uniques = max(2,random.randint(-1,3))
                
        elif self.trait['lifeform_type'].value == 'Plant-like Lifeform': 
            self.trait['species'] = Trait("Plant Species","sr(random.choice(list_of['Plants']))",summary_trait=True)
            self.trait['size'] = Trait("Lifeform Size","sr(random.choice(list_of['PlantSize']))",summary_trait=True)
            self.trait['density'] = Trait("Lifeform Density","sr(random.choice(list_of['PlantDensity']))",summary_trait=True)
            self.trait['diet'] = Trait("Lifeform Diet","sr(random.choice(list_of['PlantDiet']))")
            #self.trait['storyhook'] = Trait("Story Hook","sr(random.choice(list_of['PlantStoryHook'])")
            self.trait['reproduction'] = Trait("Reproduction Method","sr(random.choice(list_of['PlantReproMethod']))")
            self.trait['uses'] = Trait("Lifeform Uses", "sr(random.choice(list_of['PlantValue']))")
            #self.trait['demeanor'] = Trait("Lifeform Demeanor","sr(random.choice(list_of['AdvLifeformDemeanor']) ")
            self.trait['combat_profile'] = Trait("Combat Profile","sr(random.choice(list_of['BaseLifeformCombatProfile']))")
            #self.trait['reaction'] = Trait("Reaction to...","sr(random.choice(list_of['AdvLifeformReaction']) ")
            for n in range(1,max(2,random.randint(-1,3))):
                self.trait['Trait '+ str(n)] = Trait('Species Trait '+ str(n),"sr(random.choice(list_of['PlantFeatures']))")            

        for n in range(1,num_uniques):
            self.trait['Trait '+ str(n)] = Trait('Species Trait '+ str(n), "sr(random.choice(list_of['AdvLifeformUniques']))")

    def __str__(self): 
        lifeform_str=""
        for tr in self.trait:
            lifeform_str += tab(4) + str(self.trait[tr]) + "\n" 
        lifeform_str += "\n"           
        return lifeform_str    

##########################################
##  class Sentient  ######################
##########################################
       
#Return random type of sentient creature   
def get_sentient_type():
    roll = random.randint(0,100)
    
    if roll < 80:
        sentient_type = 'Standard'
    elif roll <85:
        sentient_type = 'Sentient Animal'
    elif roll <90:
        sentient_type = 'Sentient Plant'
    elif roll <95:
        sentient_type = 'Sentient Mythological'     
    else:
        sentient_type = 'Sentient Lifeform'
    return sentient_type

#Sentient Life Class - Generates Sentient Lifeform species
class SentientLife():
    
    def __init__(self,sentient_type=0,parent="self"):

        self.parent = parent
        self.object_type = "Sentient"
        self.trait = OrderedDict()

        self.trait['identity'] = Trait("Identity","'SEN-'+rand_string(3)")
        self.trait['sentient_type'] = Trait("Sentient Type","get_sentient_type()")

        if sentient_type == 0:
            sentient_type = get_sentient_type()        

        if sentient_type == "Standard":
            self.trait['race'] = Trait("Race","sr(random.choice(list_of['Races']))")            
        elif sentient_type == "Sentient Animal":
            self.trait['race'] = Trait("Race","sr(random.choice(list_of['AdvancedLifeforms']))",trait_unit="(Sentient)")
        elif sentient_type == "Sentient Plant":
            self.trait['race'] = Trait("Race","sr(random.choice(list_of['Plants']))",trait_unit="(Sentient)")
        elif sentient_type == "Sentient Mythological":
            self.trait['race'] = Trait("Race","sr(random.choice(list_of['MythologicalLifeforms']))",trait_unit="(Sentient)")
        elif sentient_type == "Sentient Lifeform":
            self.trait['race'] = Trait("Race","sr(random.choice(list_of['AdvancedLifeforms']))",trait_unit="(Sentient)") #######TRY TO ROLL RANDOM LIFEFORM AS RACE
        
        self.trait['sentient_type'].value = sentient_type
        self.trait['identity'] = Trait("Identity","'"+self.trait['race'].value+"'")
        self.trait['primary_ability_score'] = Trait("Primary ability score","sr(random.choice(list_of['DDAbilityScore']))")
        self.trait['secondary_ability_score'] = Trait("Secondary ability score","sr(random.choice(list_of['DDAbilityScore']))")
        for d in range(1,random.randint(1,4)):
            self.trait['diety'+str(d)] = Trait("Diety "+str(d),"sr(random.choice(list_of['Dieties']))") ##EXPAND TO INCLUDE PHILOSOPHIES

        self.subobject = []
    
        #Calls civilization class, generating 1-3 civilizations attached to sentient life
        for i in range(1,random.randint(2,3)): 
            self.subobject.append(Civilization(parent=self))

        #subobject = Major Cities
        #            Point of Interest
        #            Persons of interest


            
    
    def __str__(self): 
        sentient_str=""
        for tr in self.trait:
            sentient_str += tab(4) + str(self.trait[tr]) + "\n" 
        sentient_str += "\n"           
        return sentient_str      
    
    def sentient_summary(self): 
        sentient_str=""
        for tr in self.trait:
            sentient_str += tab(3) + str(self.trait[tr]) + "\n" 
        sentient_str += "\n"           
        return sentient_str    


##########################################
##  class Civilization  ##################
##########################################

#Civilization Class - Generates Civilizations for Sentient Life, and the status of that society
class Civilization():

    def __init__(self,parent="self"):
        self.parent = parent
        self.object_type = "Civilization"
        self.trait = OrderedDict()

        faction = random.randint(1,3)
        self.trait['identity'] = Trait("Identity","'CIV-'+rand_string(3)")
        self.trait['faction'] = Trait("Faction", "random.randint(1,3)")
        self.trait['faction'].value = faction
        self.trait['society_status'] = Trait("Society Status","sr(random.choice(list_of['SocietyStatus']))",trait_unit="Society")
        self.trait['era'] = Trait("Era of Time","sr(random.choice(list_of['Era']))")
        self.trait['type_of_society'] = Trait("Type of Society","sr(random.choice(list_of['TypeOfSociety']))")        
        self.trait['religion'] = Trait("Religion?","sr(random.choice(list_of['YYN']))")        

        # for p in range(1,random.randint(2,5)):
        #     self.trait['respected_profession'+str(p)] = Trait("Respected Profession "+str(p),"sr(random.choice(list_of['Jobs']))")
        self.trait['respected_professions'] = Trait("Respected Profession","get_x_from_list(random.randint(2,5),'Jobs')")
        self.trait['popular_sport'] = Trait("Popular Sport like ","get_x_from_list(2,'Sports')")
    

        self.subobject = []

        for i in range(1,random.randint(1,2)): 
            self.subobject.append(City(parent=self))

    def __str__(self):
        pass
    
    def generate(self, opt1=0,opt2=0):
        pass

#Function to get X random items from a list, Should probably make this a global function
def get_x_from_list(x,list_name):
    text_list = ""
    for l in range(1,x+1):
        text_list += sr(random.choice(list_of[list_name]))+ ", "
    return text_list
    
##########################################
##  class City  ##########################
##########################################
    
#City Class - Generates a City with a number of traits
class City():
    
    def __init__(self,parent="self"):
        self.parent = parent
        self.object_type = "City"
        self.trait = OrderedDict()

        self.trait['identity'] = Trait("Identity","'CITY-'+rand_string(3)")
        self.trait['population'] = Trait("Size by Population","sr(random.choice(list_of['CitySizes']))")
        self.trait['center_shape'] = Trait("Shape of City Center","sr(random.choice(list_of['2DShape']))")
        self.trait['outer_shape'] = Trait("Outer Shape of City","sr(random.choice(list_of['2DShape']))")
        #self.trait['fortifications']        
        self.trait['celebrities']  = Trait("Celebrities like... ","get_x_from_list(random.randint(2,4),'CharacterInspiration')")
        self.trait['food1'] = Trait("Popular Food Dish 1","sr(random.choice(list_of['Food'])) + ' Cooked by ' + sr(random.choice(list_of['WaysToCook']))")
        self.trait['food2'] = Trait("Popular Food Dish 2","sr(random.choice(list_of['Food'])) + ' Cooked by ' + sr(random.choice(list_of['WaysToCook']))")
        for t in range(1,random.randint(2,4)):
            self.trait['transit'+str(t)] = Trait("Transit Method "+str(t),"sr(random.choice(list_of['Transportation']))")
        self.trait['transit1'].value = "Path"

###########These should be Subobjects####################
        for t in range(1,random.randint(1,3)):
            self.trait['sentient_'+str(t)] = Trait("Sentient Encounter "+str(t),"sr(random.choice(list_of['Buildings']))")
            self.trait['sentient_'+str(t)+'_hook'] = Trait("Adventure Hook "+str(t),"sr(random.choice(list_of['PersonStoryHook']))")
        for t in range(1,random.randint(1,4)):
            self.trait['sentient_'+str(t)] = Trait("Point of Interest "+str(t),"sr(random.choice(list_of['Buildings']))")
            self.trait['sentient_'+str(t)+'_hook'] = Trait("Adventure Hook "+str(t),"sr(random.choice(list_of['CivilizationStoryHook']))")
        if chance(10):self.trait['wonder'] = Trait("Wonder "+str(t),"sr(random.choice(list_of['ManmadeWonders']))")
    
    def __str__(self):
        pass
    
    def generate(self, opt1=0,opt2=0):
        pass
    
##########################################
##  class Spaceship  #####################
##########################################
    
class Spaceship():
    
    def __init__(self):
        pass
    
    def __str__(self):
        pass
    
    def generate(self, opt1=0,opt2=0):
        pass
    
##########################################
##  class SpaceStation  ##################
##########################################
    
class SpaceStation():
    
    def __init__(self):
        pass
    
    def __str__(self):
        pass
    
    def generate(self, opt1=0,opt2=0):
        pass

    ### things to add ###
    # Add Moons
    # Add DC system, Consider using the Primary DC (Planet DC, Lifeform DC, Sentient DC) as the "Challenge Rating"
    #   using that DC as the DC cap for all abilities
    # Add QUESTS!
    



#################################################################################################################################################
#####  TKINTER UI  ##############################################################################################################################
################################################################################################################################################

#Tkinter UI Functions

def test_button(text):    
    print(text)

def roll_proficiency(trait):
    print(f"You rolled a {random.randint(1,trait.proficiency)}")

def lookup_link(trait):
        print("web page clicked")
        webbrowser.open("https://www.google.com/search?q="+trait.value+"&tbm=isch")

def reroll_trait(trait):
    print("In reroll_trait")
    trait.reroll()
    trait.ui_value.set(trait.value)

def change_curr_object(new_object,root="None"):
    if root != "None":
        root.destroy()
    curr_object = new_object
    load_ui(new_object)

def text_changed(var,indx,mode):
    print("TEXT HAS CHANGED TO ")

def scroll_window(event,canvas):
    canvas.configure(scrollregion=canvas.bbox("all"),width=0.32*screenX,height=screenY)

def reroll_object(object,parent,root="None"):
    if root != "None":
        root.destroy()
    object.__init__(parent=parent)
    load_ui(object)

#######################
#UI is generated HERE
#######################
def load_ui(new_object):
    curr_object = new_object

    root = Toplevel()
    root.geometry(str(screenX)+"x"+str(screenY))
    root.title("Space DD sub")

    leftFrame = Frame(root,bg = "yellow",width=0.32*screenX,height = screenY)
    rightFrame = Frame(root, bg = "black",width=0.68*screenX,height = screenY)

    leftFrame.grid(row=0,column=0,sticky="NESW")
    rightFrame.grid(row=0,column=1,sticky="NESW")


    # subobjectsFramePrime = Frame(leftFrame,width=0.32*screenX)
    # subobjectsFramePrime.grid(row=1,column=0,padx=2,pady=2,sticky="NESW")
    subobjectsCanvas = Canvas(leftFrame) #,bg = "pink",width=0.32*screenX
    subobjectsCanvas.pack(side="left", fill = "both")

    subobjectsFrame = Frame(subobjectsCanvas)
    subobjectsFrame.pack(fill = "both")

    traitsFrame = Frame(subobjectsFrame,bg = "orange",width=0.32*screenX)
    traitsFrame.grid(row=0,column=0,padx=0,pady=0,sticky="NEW")

    subobjectsCanvas.create_window((0,0),window=subobjectsFrame,anchor='nw')
    subobjectsFrame.bind("<Configure>",lambda event, canvas = subobjectsCanvas: scroll_window(event, canvas))

    scrollbar = Scrollbar(leftFrame,orient="vertical",command=subobjectsCanvas.yview)
    scrollbar.pack(side="right", fill="y")

    subobjectsCanvas.configure(yscrollcommand=scrollbar.set)


    objectLabel = Label(traitsFrame, text = str(curr_object.trait['identity'].value))
    objectLabel.grid(row=0,columnspan=5,sticky="W")

    trait_counter = 1    

    if (curr_object.parent != "self"):
        ui_collapse_button = Button(traitsFrame,text="Collapse", command= lambda new_object = curr_object.parent, root = root: change_curr_object(new_object,root))
        ui_collapse_button.grid(row=trait_counter,column=0,padx=1,columnspan=3,sticky=W)
        trait_counter += 1
    ui_reroll_object_button = Button(traitsFrame,text="Reroll", command= lambda curr_object = curr_object, root = root, parent = curr_object.parent: reroll_object(curr_object,parent,root))
    ui_reroll_object_button.grid(row=trait_counter,column=5,padx=1,columnspan=2,sticky=E)
    trait_counter += 1

    #### GENERATE UI FOR TRAITS #######
    for tr in curr_object.trait: 
        curr_object.trait[tr].generate_trait_ui(traitsFrame,trait_counter)
        trait_counter += 1

    #### GENERATE UI FOR SUB-OBJECTS #####
    try:
        subobject_counter=1
        for so in curr_object.subobject:            
            subobjectframe = LabelFrame(subobjectsFrame, text = str(so.trait['identity'].value))        
            subobjectframe.grid(row=subobject_counter, column = 0, columnspan = 8,sticky="NESW") #, columnspan = 8
            subobject_counter += 1

            so_row_counter = 0
            tr_button = Button(subobjectframe,text="Expand",command = lambda new_object = so, root = root: change_curr_object(new_object,root))
            tr_button.grid(row=so_row_counter,column=0,columnspan=3,sticky="W")
            so_row_counter += 1
            for tr in so.trait:
                if(so.trait[tr].summary_trait == True):
                    so.trait[tr].generate_trait_ui(subobjectframe,so_row_counter)
                    so_row_counter += 1
    except AttributeError:
        pass

    canvas_width = 0.68*screenX
    canvas_height = 0.7*screenY-2
    canvas_scaleX = 1
    canvas_scaleY = 1
    viewCanvas = Canvas(rightFrame,bg = "black", width=canvas_width,height = canvas_height)
    viewCanvas.grid(row=0,column=0,padx=0,pady=0,sticky="NESW")



    if curr_object.object_type == "Sol":
        #print ("Creating Sol UI")
        canvas_scaleX = canvas_width / (curr_object.trait['edge_helio'].value*1.1)
        canvas_scaleY = canvas_height / (curr_object.trait['edge_helio'].value*1.1)
        create_oval_by_center(viewCanvas,canvas_width/2,canvas_height/2,float(curr_object.trait['star_diameter'].value*5+5),color="yellow")
        for so in curr_object.subobject:
            create_object_with_orbit(viewCanvas,so,canvas_width,canvas_height,canvas_scaleX,canvas_scaleY)

            

    elif curr_object.object_type == "Planet":
        print ("Creating Planet UI")
    else:
        print ("Creating Other UI")


    journalFrame=Frame(rightFrame,bg = "white", width=0.68*screenX,height = 0.2*screenY)
    buttonFrame=Frame(rightFrame,bg = "red",width=0.68*screenX,height = 0.1*screenY)

    journalFrame.grid(row=1,column=0,padx=0,pady=0,sticky="NESW")
    buttonFrame.grid(row=2,column=0,padx=0,pady=0,sticky="NESW")

    root.mainloop()

#Creates oval by center, for Planet Orbits
def create_oval_by_center(canvas,x,y,radius,color="white",solid=True):
    if solid:
        canvas.create_oval(x-radius,y-radius,x+radius,y+radius, fill = color)
    else:
        canvas.create_oval(x-radius,y-radius,x+radius,y+radius, outline = color)

#creates circle for planet, with its orbit
def create_object_with_orbit(canvas,planet,canvasX,canvasY,canvas_scaleX,canvas_scaleY):
    orbit = int(planet.trait['orbit_radius'].value)
    size = planet.trait['size'].value+2
    x1,y1 = canvasX/2-orbit*canvas_scaleX, canvasY/2-orbit*canvas_scaleY
    x2,y2 = canvasX/2+orbit*canvas_scaleX,canvasY/2+orbit*canvas_scaleY
    canvas.create_oval(x1,y1,x2,y2,outline = "white")

    objectX=math.cos(planet.trait['orbit_radians'].value)*orbit
    objectY=math.sin(planet.trait['orbit_radians'].value)*orbit
    x1,y1 = canvasX/2-objectX*canvas_scaleX-size/2,canvasY/2-objectY*canvas_scaleY-size/2
    x2,y2 = canvasX/2-objectX*canvas_scaleX+size/2,canvasY/2-objectY*canvas_scaleY+size/2
    canvas.create_oval(x1,y1,x2,y2,fill=planet.trait['color'].value)
    canvas.create_text(x2,y2, anchor = "nw",fill="white", text=planet.trait['identity'].value)


################################
#Beginning of Program
################################

#Import excel file
import_random_lists_from_file()

#Initialize a Solar System
sol = {}

def new_campaign_button():
    pass

def load_campaign_button():
    #Open a file for editing
    filepath = askopenfilename(
        filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
    )
    if not filepath:
        return
    #txt_edit.delete(1.0, tk.END)
    with open(filepath, "r") as input_file:
        text = input_file.read()
        txt_edit.insert(tk.END, text)
    window.title(f"Load Campaign - {filepath}")
    pass

#Generate a Sol and generate UI
def new_galaxy_button():
    sol=Sol()
    change_curr_object(sol)


#generate random board UI
def random_board_button():
    board = Toplevel()
    board.geometry(str(screenX)+"x"+str(screenY))
    board.title("Random Generator") 

    list_keys = list(list_of.keys())
   
    for y in range(50):
        for x in range(10):  
            this_key = list_keys.pop(0)
            b=Button(board,text= this_key, command = lambda this_key = this_key: rando(this_key))   
            b.grid(column = x,row = y, sticky='nesw', padx = 1, pady = 1) 
            #button[]

        

main_menu = Tk()
main_menu.geometry("200x200")
main_menu.title("Space DD - Main Menu")
new_campaign_button = Button(main_menu,text="New Campaign", command= new_campaign_button)
new_campaign_button.pack()
load_campaign_button = Button(main_menu,text="Load Campaign", command= load_campaign_button)
load_campaign_button.pack()
new_galaxy_button = Button(main_menu,text="New Galaxy", command= new_galaxy_button)
new_galaxy_button.pack()
random_board_button = Button(main_menu,text="Random Board", command= random_board_button)
random_board_button.pack()

############## IMAGES FOR UI##################
image_d6 = PhotoImage(file = r"d6-15x15.png") 
image_reroll = PhotoImage(file = r"reroll.png") 
image_lookup = PhotoImage(file = r"lookup.png") 
image_journal = PhotoImage(file = r"journal.png") 


main_menu.mainloop()

# for i in range(1,5):
#     text = sr(random.choice(list_of['NestTest']))
#     print(text)